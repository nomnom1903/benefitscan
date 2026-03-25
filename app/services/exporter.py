"""
app/services/exporter.py — Stage 5: Excel comparison spreadsheet generation

Produces a professional .xlsx file the broker can send directly to their client.
Two sheets:
  1. "Plan Comparison"     — one row per plan, one column per field, formatted
  2. "Extraction Summary"  — metadata about when/how each plan was extracted

Why openpyxl over xlsxwriter:
openpyxl can both read and write .xlsx files. xlsxwriter is write-only.
We don't need read capability today, but we might in V2 (e.g. updating a
previously exported file). Also, openpyxl's API maps more directly to how
brokers think about Excel — cells, rows, columns — vs. xlsxwriter's more
programmatic approach.
"""

import json
import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, Protection
)
from openpyxl.utils import get_column_letter

from app.models.sbc import SBC_FIELD_KEYS, SBC_FIELD_LABELS

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# Color palette — matches the frontend UI
# ─────────────────────────────────────────────────────────────────────────────
COLOR_NAVY = "0F172A"          # Header background
COLOR_WHITE = "FFFFFF"         # Header text / default cell background
COLOR_LIGHT_GRAY = "F8FAFC"    # Alternating row background
COLOR_GREEN_LIGHT = "DCFCE7"   # OK cell background (subtle)
COLOR_YELLOW_LIGHT = "FEF9C3"  # Review/flagged cell background
COLOR_RED_LIGHT = "FEE2E2"     # Missing cell background
COLOR_AMBER = "D97706"         # Accent (used for non-compliant)
COLOR_AMBER_LIGHT = "FEF3C7"   # Non-compliant cell background


def export_to_excel(plans: list[dict], output_dir: Path) -> Path:
    """
    Generate a formatted Excel comparison spreadsheet from a list of extracted plans.

    Args:
        plans: list of plan dicts (from SBCPlanDB.to_dict()), one per SBC
        output_dir: directory to save the .xlsx file

    Returns:
        Path to the generated .xlsx file

    Raises:
        ValueError: if plans list is empty
    """
    if not plans:
        raise ValueError("Cannot export: no plans provided")

    wb = openpyxl.Workbook()

    # --- Sheet 1: Plan Comparison ---
    _build_comparison_sheet(wb.active, plans)
    wb.active.title = "Plan Comparison"

    # --- Sheet 2: Extraction Summary ---
    summary_sheet = wb.create_sheet("Extraction Summary")
    _build_summary_sheet(summary_sheet, plans)

    # Save to disk with a timestamped filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"BenefitScan_Comparison_{timestamp}.xlsx"
    output_path = output_dir / filename

    wb.save(str(output_path))
    logger.info(f"Excel export saved: {output_path}")
    return output_path


def _build_comparison_sheet(ws, plans: list[dict]) -> None:
    """
    Build the Plan Comparison sheet.

    Layout:
      Row 1: Column headers (frozen, navy background)
      Row 2+: One row per plan (alternating light gray / white)
      Column A: Plan name (frozen so it stays visible while scrolling right)
    """
    # ─── Header row ───────────────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, color=COLOR_WHITE, size=10)
    header_fill = PatternFill("solid", fgColor=COLOR_NAVY)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["Plan Name"] + [
        SBC_FIELD_LABELS.get(key, key.replace("_", " ").title())
        for key in SBC_FIELD_KEYS
        if key != "plan_name"  # plan_name is first column, skip in the loop
    ]
    # Rebuild full ordered column list: plan_name first, then the rest
    ordered_keys = ["plan_name"] + [k for k in SBC_FIELD_KEYS if k != "plan_name"]
    ordered_headers = [SBC_FIELD_LABELS.get(k, k.replace("_", " ").title()) for k in ordered_keys]

    for col_idx, header_text in enumerate(ordered_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # ─── Data rows ────────────────────────────────────────────────────────────
    data_font = Font(name="Calibri", size=10)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for row_idx, plan in enumerate(plans, start=2):
        # Alternate row background for readability
        row_fill_color = COLOR_LIGHT_GRAY if row_idx % 2 == 0 else COLOR_WHITE
        default_fill = PatternFill("solid", fgColor=row_fill_color)

        # Parse the validation report for this plan (may be None if extraction failed)
        validation = plan.get("validation_report") or {}
        field_results = validation.get("field_results", {})

        for col_idx, field_key in enumerate(ordered_keys, start=1):
            value = plan.get(field_key)
            cell = ws.cell(row=row_idx, column=col_idx, value=value or "")

            cell.font = data_font
            cell.alignment = left_alignment if col_idx == 1 else center_alignment

            # Color-code the cell based on validation status
            field_status = field_results.get(field_key, {}).get("status", "OK")
            cell.fill = _status_to_fill(field_status, default_fill)

    # ─── Freeze panes ─────────────────────────────────────────────────────────
    # Freeze row 1 (header) AND column A (plan name).
    # Cell B2 is the "corner" — everything above and to the left stays fixed.
    ws.freeze_panes = "B2"

    # ─── Column widths ────────────────────────────────────────────────────────
    # Column A (plan name) gets a fixed wide width. Others auto-sized.
    ws.column_dimensions["A"].width = 28
    for col_idx in range(2, len(ordered_keys) + 1):
        col_letter = get_column_letter(col_idx)
        # Auto-width: scan all cell values and set width to longest + padding
        max_length = len(ordered_headers[col_idx - 1])  # at minimum the header length
        for row_idx in range(2, len(plans) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 25)  # cap at 25

    # ─── Row height ───────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 36  # Taller header row for wrapped text
    for row_idx in range(2, len(plans) + 2):
        ws.row_dimensions[row_idx].height = 20


def _build_summary_sheet(ws, plans: list[dict]) -> None:
    """
    Build the Extraction Summary sheet.
    Shows metadata about each extraction: filename, timestamp, field counts.
    """
    header_font = Font(name="Calibri", bold=True, color=COLOR_WHITE, size=10)
    header_fill = PatternFill("solid", fgColor=COLOR_NAVY)

    headers = [
        "Plan Name", "Source File", "Extracted At",
        "Total Fields", "OK", "Missing", "Review", "Non-Compliant", "Overall Status"
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, plan in enumerate(plans, start=2):
        validation = plan.get("validation_report") or {}
        summary = validation.get("summary", {})
        overall = validation.get("overall_status", "—")

        row_data = [
            plan.get("plan_name") or "Unknown Plan",
            plan.get("upload_filename", ""),
            plan.get("extracted_at", ""),
            summary.get("total_fields", 0),
            summary.get("ok_count", 0),
            summary.get("missing_count", 0),
            summary.get("review_count", 0),
            summary.get("non_compliant_count", 0),
            overall,
        ]
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-size columns
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20

    ws.freeze_panes = "A2"


def _status_to_fill(status: str, default_fill: PatternFill) -> PatternFill:
    """Return the appropriate cell background fill for a validation status."""
    fills = {
        "OK": default_fill,                                       # subtle row alternation
        "Missing": PatternFill("solid", fgColor=COLOR_RED_LIGHT),
        "Review": PatternFill("solid", fgColor=COLOR_YELLOW_LIGHT),
        "Non-Compliant": PatternFill("solid", fgColor=COLOR_AMBER_LIGHT),
    }
    return fills.get(status, default_fill)
